#!/usr/bin/env python3
"""Ingest official Catalyst project tracking spreadsheet into PROOF database.

Source: https://docs.google.com/spreadsheets/d/1bfnWFa94Y7Zj0G7dtpo9W1nAYGovJbswipxiHT4UE3g

This spreadsheet is maintained by the Catalyst team and contains:
- Fund 2-14 project reporting sheets (status, amounts, proposer, challenge, reporting compliance)
- COMPLETED sheet with completion dates
- Monthly TXs with on-chain distribution transaction hashes
- Summary stats per fund

What this script enriches:
1. Official project status (COMPLETED / DID NOT FINISH / IN PROGRESS)
2. Completion dates from COMPLETED sheet
3. Reporting compliance scores (% of reports submitted)
4. Verified financial amounts (requested, distributed, remaining)
5. Spreadsheet numeric ID stored as ideascaleId for future cross-referencing

Usage:
    python ingest_official_spreadsheet.py [--dry-run] [--xlsx path/to/file.xlsx]
"""

from __future__ import annotations

import argparse
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
import psycopg2
import psycopg2.extras


# Map spreadsheet fund names to fund numbers
FUND_SHEET_MAP = {
    "Fund2 Reporting": 2,
    "Fund3 Reporting": 3,
    "Fund4 Reporting": 4,
    "Fund5 Reporting": 5,
    "Fund6 Reporting": 6,
    "Fund7 Reporting": 7,
    "Fund8 Reporting": 8,
    "Fund9 Reporting": 9,
    "Fund10 Reporting": 10,
    "Fund11 Reporting": 11,
    "Fund12 Reporting": 12,
    "Fund13 Reporting": 13,
    "Fund14 Reporting": 14,
}

# Map spreadsheet status to our status/fundingStatus
STATUS_MAP = {
    "COMPLETED": ("complete", "funded"),
    "DID NOT FINISH": ("abandoned", "funded"),
    "IN PROGRESS": ("in_progress", "funded"),
    "PENDING": ("in_progress", "funded"),
    "PAUSED": ("paused", "funded"),
    "ON HOLD": ("paused", "funded"),
    "SUSPENDED": ("suspended", "funded"),
    "TERMINATED": ("terminated", "funded"),
}

# Header row index (0-based) — row 4 in all fund sheets
HEADER_ROW = 4
# First data row (0-based) — row 8 in all fund sheets (after header, totals, reporting %, count)
DATA_START_ROW = 8


def load_env(env_path: str) -> None:
    """Load .env file into os.environ."""
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                os.environ[k] = v.strip('"').strip("'")


def parse_amount(val) -> float | None:
    """Parse a dollar amount from the spreadsheet."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).replace("$", "").replace(",", "").strip()
    if not s or s == "0":
        return 0.0
    try:
        return float(s)
    except ValueError:
        return None


def count_reports(row_values: list, report_start_idx: int) -> tuple[int, int]:
    """Count submitted vs total report slots from TRUE/FALSE columns.
    
    Returns (submitted, total) where submitted = count of TRUE values
    and total = count of non-empty report cells.
    """
    submitted = 0
    total = 0
    for val in row_values[report_start_idx:]:
        if val is None or str(val).strip() == "":
            continue
        s = str(val).strip().upper()
        if s in ("COMPLETED", "PAUSED", "REACTIVATED"):
            continue
        total += 1
        if s == "TRUE" or s == "1":
            submitted += 1
    return submitted, total


def parse_fund_sheet(ws, fund_number: int) -> list[dict]:
    """Parse a fund reporting sheet and return list of project dicts."""
    projects = []
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < DATA_START_ROW + 1:
        return projects

    # Find the header row to identify column positions
    header = rows[HEADER_ROW]
    
    # Find the index of the first reporting column (after Challenge)
    # Columns: #, ID, Project Name, Project Status, Requested, Distributed, Remaining, Proposer, Challenge, ...reports
    report_start_idx = 9  # Default: column J onwards

    for i in range(DATA_START_ROW, len(rows)):
        row = rows[i]
        if row is None or len(row) < 8:
            continue

        # Column B = ID (numeric, like 500001)
        proj_id = row[1]
        if proj_id is None:
            continue

        # Skip non-numeric IDs (could be subtotals, etc.)
        try:
            proj_id = int(float(str(proj_id)))
        except (ValueError, TypeError):
            continue

        title = str(row[2]).strip() if row[2] else None
        if not title:
            continue

        status_raw = str(row[3]).strip().upper() if row[3] else None
        requested = parse_amount(row[4])
        distributed = parse_amount(row[5])
        remaining = parse_amount(row[6])
        proposer = str(row[7]).strip() if row[7] else None
        challenge = str(row[8]).strip() if row[8] else None

        # Count reporting compliance
        submitted, total = count_reports(list(row), report_start_idx)

        # Map status
        our_status, our_funding_status = STATUS_MAP.get(
            status_raw, ("in_progress", "funded")
        )

        projects.append({
            "spreadsheet_id": proj_id,
            "fund_number": fund_number,
            "title": title,
            "status_raw": status_raw,
            "status": our_status,
            "funding_status": our_funding_status,
            "requested": requested,
            "distributed": distributed,
            "remaining": remaining,
            "proposer": proposer,
            "challenge": challenge,
            "reports_submitted": submitted,
            "reports_total": total,
            "reporting_compliance": round(submitted / total, 4) if total > 0 else None,
        })

    return projects


def parse_completed_sheet(ws) -> dict[str, datetime]:
    """Parse COMPLETED sheet to get completion dates.
    
    Returns dict mapping lowercase title -> completion datetime.
    """
    completion_dates = {}
    rows = list(ws.iter_rows(values_only=True))

    # Header is at row 5: '', '#', 'Project ID#', 'Project Name', 'Project Status',
    # 'Requested', 'Distributed', 'Remaining', 'Proposer', 'Fund #', 'Challenge', 'Date Completed'
    for i in range(7, len(rows)):  # Data starts at row 7
        row = rows[i]
        if row is None or len(row) < 12:
            continue

        title = row[3]
        date_completed = row[11]

        if title is None or date_completed is None:
            continue

        title = str(title).strip()
        if not title:
            continue

        if isinstance(date_completed, datetime):
            completion_dates[title.lower()] = date_completed
        elif isinstance(date_completed, str):
            try:
                completion_dates[title.lower()] = datetime.strptime(
                    date_completed.strip(), "%Y-%m-%d"
                )
            except ValueError:
                pass

    return completion_dates


def parse_monthly_txs(ws) -> list[dict]:
    """Parse Monthly TXs sheet to get distribution transaction hashes.
    
    Returns list of dicts with date, fund_number, tx_hash.
    """
    txs = []
    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        return txs

    # Header row 0: '', 'Distribution #', 'Actual date', 'Telegram Announcement',
    # 'Fund4 (tx id)', 'Fund5 (tx id)', 'Fund6 (tx id)', ...
    header = rows[0]
    fund_cols = {}
    for col_idx, val in enumerate(header):
        if val and "Fund" in str(val) and "tx id" in str(val):
            # Extract fund number from "Fund4 (tx id)"
            fund_num = "".join(c for c in str(val).split("(")[0] if c.isdigit())
            if fund_num:
                fund_cols[col_idx] = int(fund_num)

    for i in range(1, len(rows)):
        row = rows[i]
        if row is None:
            continue

        tx_date = row[2] if len(row) > 2 else None
        if tx_date is None:
            continue

        if isinstance(tx_date, str):
            try:
                tx_date = datetime.strptime(tx_date.strip(), "%Y-%m-%d")
            except ValueError:
                continue

        for col_idx, fund_num in fund_cols.items():
            if col_idx < len(row) and row[col_idx]:
                tx_hash = str(row[col_idx]).strip()
                if len(tx_hash) > 20:  # Likely a real tx hash
                    txs.append({
                        "date": tx_date,
                        "fund_number": fund_num,
                        "tx_hash": tx_hash,
                    })

    return txs


def update_database(
    conn,
    all_projects: list[dict],
    completion_dates: dict[str, datetime],
    dry_run: bool = False,
) -> dict:
    """Match spreadsheet projects to DB and update records."""
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    stats = {
        "matched": 0,
        "not_found": 0,
        "updated_status": 0,
        "updated_completion": 0,
        "updated_amounts": 0,
        "updated_ideascale_id": 0,
        "skipped_multi_match": 0,
        "total": len(all_projects),
    }

    # Get fund IDs
    cur.execute('SELECT id, number FROM "Fund"')
    fund_map = {row["number"]: row["id"] for row in cur.fetchall()}

    for proj in all_projects:
        fund_id = fund_map.get(proj["fund_number"])
        if not fund_id:
            stats["not_found"] += 1
            continue

        # Match by title + fund
        cur.execute(
            """
            SELECT id, title, status, "fundingStatus", "completedAt",
                   "fundingAmount", "amountReceived", "amountRemaining",
                   "ideascaleId"
            FROM "Project"
            WHERE LOWER(title) = LOWER(%s) AND "fundId" = %s
            """,
            (proj["title"], fund_id),
        )
        rows = cur.fetchall()

        if len(rows) == 0:
            stats["not_found"] += 1
            continue
        if len(rows) > 1:
            stats["skipped_multi_match"] += 1
            continue

        db_row = rows[0]
        stats["matched"] += 1
        updates = []
        params = []

        # 1. Update status if different
        if proj["status"] and db_row["status"] != proj["status"]:
            updates.append('"status" = %s')
            params.append(proj["status"])
            stats["updated_status"] += 1

        # 2. Update completedAt from COMPLETED sheet
        completion_date = completion_dates.get(proj["title"].lower())
        if completion_date and db_row["completedAt"] is None:
            updates.append('"completedAt" = %s')
            params.append(completion_date)
            stats["updated_completion"] += 1

        # 3. Update ideascaleId if not set
        if db_row["ideascaleId"] is None and proj["spreadsheet_id"]:
            updates.append('"ideascaleId" = %s')
            params.append(str(proj["spreadsheet_id"]))
            stats["updated_ideascale_id"] += 1

        # 4. Verify amounts — update amountReceived if distributed differs
        if proj["distributed"] is not None:
            db_received = float(db_row["amountReceived"] or 0)
            if abs(db_received - proj["distributed"]) > 1.0:
                updates.append('"amountReceived" = %s')
                params.append(proj["distributed"])
                updates.append('"amountRemaining" = %s')
                params.append(proj["remaining"] or 0)
                stats["updated_amounts"] += 1

        if updates and not dry_run:
            updates.append('"updatedAt" = NOW()')
            sql = f'UPDATE "Project" SET {", ".join(updates)} WHERE id = %s'
            params.append(db_row["id"])
            cur.execute(sql, params)

    if not dry_run:
        conn.commit()

    cur.close()
    return stats


def main() -> None:
    ap = argparse.ArgumentParser(description="Ingest official Catalyst spreadsheet")
    ap.add_argument(
        "--xlsx",
        default=str(Path(__file__).parent / "output" / "catalyst_official.xlsx"),
        help="Path to the downloaded xlsx file",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and match but don't write to database",
    )
    ap.add_argument(
        "--env",
        default=str(Path(__file__).parents[2] / ".env"),
        help="Path to .env file",
    )
    args = ap.parse_args()

    # Load environment
    if os.path.exists(args.env):
        load_env(args.env)

    dsn = os.environ.get("DATABASE_URL")
    if not dsn:
        raise SystemExit("Missing DATABASE_URL")

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise SystemExit(f"Missing xlsx file: {xlsx_path}")

    print(f"Loading {xlsx_path} ...")
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # 1. Parse all fund reporting sheets
    all_projects = []
    for sheet_name, fund_number in FUND_SHEET_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP: {sheet_name} not found")
            continue
        ws = wb[sheet_name]
        projects = parse_fund_sheet(ws, fund_number)
        print(f"  {sheet_name}: {len(projects)} projects")
        all_projects.extend(projects)

    print(f"\nTotal projects from fund sheets: {len(all_projects)}")

    # 2. Parse COMPLETED sheet for completion dates
    completion_dates = {}
    if "COMPLETED" in wb.sheetnames:
        completion_dates = parse_completed_sheet(wb["COMPLETED"])
        print(f"Completion dates found: {len(completion_dates)}")

    # 3. Parse Monthly TXs for distribution tx hashes
    monthly_txs = []
    if "Monthly TXs" in wb.sheetnames:
        monthly_txs = parse_monthly_txs(wb["Monthly TXs"])
        print(f"Distribution tx hashes found: {len(monthly_txs)}")

    wb.close()

    # Print status distribution
    status_counts = {}
    for p in all_projects:
        s = p["status_raw"] or "UNKNOWN"
        status_counts[s] = status_counts.get(s, 0) + 1
    print(f"\nSpreadsheet status distribution:")
    for s, c in sorted(status_counts.items(), key=lambda x: -x[1]):
        print(f"  {s}: {c}")

    # Print reporting compliance summary
    compliant = [p for p in all_projects if p["reporting_compliance"] is not None]
    if compliant:
        avg = sum(p["reporting_compliance"] for p in compliant) / len(compliant)
        print(f"\nReporting compliance: avg={avg:.1%} across {len(compliant)} projects")

    # 4. Connect to database and update
    print(f"\n{'DRY RUN — ' if args.dry_run else ''}Connecting to database...")
    conn = psycopg2.connect(dsn)

    try:
        stats = update_database(conn, all_projects, completion_dates, dry_run=args.dry_run)

        print(f"\n=== Results ===")
        print(f"  Total spreadsheet projects: {stats['total']}")
        print(f"  Matched to DB:              {stats['matched']}")
        print(f"  Not found in DB:            {stats['not_found']}")
        print(f"  Multi-match (skipped):       {stats['skipped_multi_match']}")
        print(f"  Status updated:             {stats['updated_status']}")
        print(f"  Completion date set:        {stats['updated_completion']}")
        print(f"  Amounts corrected:          {stats['updated_amounts']}")
        print(f"  IdeascaleId set:            {stats['updated_ideascale_id']}")

        if args.dry_run:
            print("\n(Dry run — no changes written)")
    finally:
        conn.close()

    print("\nDone.")


if __name__ == "__main__":
    main()
